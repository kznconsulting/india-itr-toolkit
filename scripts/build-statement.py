#!/usr/bin/env python3
"""Generic ITR computation-statement builder (works for any client).

Reads a per-client, per-year statement-data.json (template: templates/
statement-data-template.json; how to fill it from AIS/TIS/26AS PDFs:
docs/extract-from-pdfs.md) and writes the operator's statement Excel:
an Income sheet plus Dividends, Interest income, Capital Gains, and a
LT-capital-loss roll-forward sheet, mirroring the practice's house format.

Usage:
    python3 scripts/build-statement.py <statement-data.json> [--out <file.xlsx>]
    (or: bun run statement <statement-data.json>)

Default output: "<client.displayName>_AY <yy-yy>_Statement.xlsx" next to the
data file.

Design notes (read before editing):
- Tax slab rules live ONLY in scripts/lib/tax.ts. This script fetches them as
  JSON via `bun scripts/tax-cli.ts` so the Excel and the JSON validator can
  never disagree about rates.
- Cells carry real formulas (SUM, band-wise slab terms, 87A rebate with
  marginal relief, cess). After saving, this script re-evaluates every formula
  it wrote and asserts the results against (a) the reconciliation targets in
  the data file (TIS totals) and (b) scripts/lib/tax.ts's computed tax. A
  build that does not reconcile exits non-zero naming the mismatch, so a
  green run means the sheet's arithmetic is verified even without opening it.
- Special-rate capital gains ARE modelled: STCG u/s 111A and LTCG u/s 112A
  (with its Rs. 1,25,000 exemption) are priced by scripts/lib/tax.ts's
  computeTotalTax per regime column - kept out of the slab-rate/
  deduction-eligible base (no Chapter VI-A set-off against them), with the
  resident-only basic-exemption adjustment and the resident-only 87A gate
  (client.residentialStatus: "NRI"/"Non-Resident" suppresses the rebate).
- Deliberate guards (extend consciously, do not work around):
    * aborts if TOTAL income (incl. special-rate gains) exceeds Rs. 50 lakh
      (surcharge not modelled in the sheet formulas)
    * aborts on non-equity capital-gains items (assetClass other than
      "equity"): s.112 property/debt/gold rates are not modelled
"""

import json
import re
import subprocess
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

REPO_ROOT = Path(__file__).resolve().parent.parent

ACC = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"??_);_(@_)'
BASE = Font(name="Calibri", size=11)
BOLD = Font(name="Calibri", size=11, bold=True)
NAME_FONT = Font(name="Verdana", size=9, bold=True)
NOTE = Font(name="Calibri", size=10, italic=True, color="666666")

SURCHARGE_GUARD = 5_000_000


def die(msg):
    print(f"ERROR: {msg}", file=sys.stderr)
    sys.exit(1)


# Toolkit gaps are not data errors: retrying, tweaking inputs, or working around
# them burns tokens and risks a wrong return. One instruction, then move on.
PARK = (
    "\nThis is a TOOLKIT GAP, not a data error - do NOT retry or work around it. "
    "On an operator machine: git pull, add an OPEN entry to docs/missing-functionality.md, "
    "commit+push, tell the operator this client is PARKED pending a toolkit update, and "
    "move to the next client (AGENTS.md 'Escalate, don't extend')."
)


def die_gap(msg):
    die(msg + PARK)


def tax_cli(*args):
    """Call bun scripts/tax-cli.ts and return parsed JSON."""
    cmd = ["bun", "scripts/tax-cli.ts", *args]
    try:
        out = subprocess.run(cmd, cwd=REPO_ROOT, capture_output=True, text=True, check=True)
    except FileNotFoundError:
        die("bun not found on PATH (needed for tax rules) - see docs/porting.md")
    except subprocess.CalledProcessError as e:
        die(f"tax-cli failed: {e.stdout or e.stderr}")
    result = json.loads(out.stdout)
    if "error" in result:
        die(f"tax-cli: {result['error']}")
    return result


# ---------------------------------------------------------------- helpers
class Sheet:
    """Thin wrapper: cell writes with house styles + a row cursor."""

    def __init__(self, ws):
        self.ws = ws
        self.r = 1

    def put(self, coord, value, font=BASE, fmt=None):
        c = self.ws[coord]
        c.value = value
        c.font = font
        if fmt:
            c.number_format = fmt
        return c

    def amt(self, coord, value, bold=False):
        return self.put(coord, value, BOLD if bold else BASE, ACC)

    def row(self, cells, bold=False, note=False):
        """Write {col: value} on the current row; amounts detected by int/float or leading '='."""
        font = NOTE if note else (BOLD if bold else BASE)
        for col, v in cells.items():
            if isinstance(v, (int, float)) or (isinstance(v, str) and v.startswith("=")):
                self.put(f"{col}{self.r}", v, BOLD if bold else BASE, ACC)
            else:
                self.put(f"{col}{self.r}", v, font)
        r = self.r
        self.r += 1
        return r

    def skip(self, n=1):
        self.r += n

    def widths(self, spec):
        for col, w in spec.items():
            self.ws.column_dimensions[col].width = w


def fmt_rate(rate):
    return f"{rate:g}"


def slab_formula(ti_cell, slabs):
    """Band-wise progressive tax as an explicit Excel formula (correct for any TI)."""
    terms = []
    lower = 0
    for band in slabs:
        upper, rate = band.get("upTo"), band["rate"]
        if rate > 0:
            if upper is None:
                terms.append(f"MAX(0,{ti_cell}-{lower})*{fmt_rate(rate)}")
            elif upper > lower:
                terms.append(f"MAX(0,MIN({ti_cell},{upper})-{lower})*{fmt_rate(rate)}")
        if upper is not None:
            lower = max(lower, upper)
    return "=ROUND(" + "+".join(terms) + ",0)"


def rebate_formula(ti_cell, tax_cell, rebate):
    thr, cap = rebate["threshold"], rebate["cap"]
    if rebate["marginalRelief"]:
        return f"=-IF({ti_cell}<={thr},MIN({tax_cell},{cap}),MAX(0,{tax_cell}-({ti_cell}-{thr})))"
    return f"=-IF({ti_cell}<={thr},MIN({tax_cell},{cap}),0)"


def py_band_tax(slabs, ti):
    tax, lower = 0, 0
    for band in slabs:
        upper = band.get("upTo") or float("inf")
        if upper > lower and ti > lower:
            tax += (min(ti, upper) - lower) * band["rate"]
        lower = max(lower, upper) if upper != float("inf") else lower
    return round(tax)


def ay_short(ay):  # "2026-27" -> "26-27"
    return ay[2:]


# ---------------------------------------------------------------- build
def build(data, out_path, data_dir=None):
    client = data["client"]
    ay = data["assessmentYear"]
    prior = data.get("priorYearLabel", "")
    short = ay_short(ay)

    age = tax_cli("ageband", "--dob", client["dob"], "--ay", ay)["age"]
    res_raw = str(client.get("residentialStatus", "resident")).strip().lower()
    res = "nri" if ("nri" in res_raw or "non-resident" in res_raw or res_raw == "nr") else "resident"
    rules_new = tax_cli("rules", "--regime", "new", "--ay", ay, "--age", age)
    # old-regime senior/super-senior exemption slabs are resident-only
    rules_old = tax_cli("rules", "--regime", "old", "--ay", ay,
                        "--age", "lt60" if res == "nri" else age)

    # ---- derived figures (python side, later asserted against the sheet)
    biz_items = data.get("business", {}).get("items", [])
    biz_gross = sum(b["gross"] for b in biz_items)
    biz_tds = sum(b.get("tds", 0) for b in biz_items)
    presumptive = sum(round(b["gross"] * b["presumptiveRate"]) for b in biz_items)

    salary = data.get("salary", {})
    salary_gross = salary.get("gross", 0) or 0
    salary_tds = salary.get("tds", 0) or 0
    std_ded_old = rules_old["standardDeduction"] if salary_gross else 0
    std_ded_new = rules_new["standardDeduction"] if salary_gross else 0

    interest = data.get("interest", {})
    sb_items = interest.get("savings", [])
    dep_items = interest.get("deposits", [])
    other_items = interest.get("other", [])
    sb_total = sum(i["amount"] for i in sb_items)
    # residents never see savings-account TDS (194A exempts savings interest), but an
    # NRI's savings interest carries flat s.195 TDS like any other NRI income stream
    sb_tds = interest.get("savingsTds", 0)
    dep_total = sum(i["amount"] for i in dep_items)
    dep_tds = interest.get("depositsTds", 0)
    other_total = sum(i["amount"] for i in other_items)
    other_tds = interest.get("otherTds", 0)
    refund_int = (interest.get("refundInterest") or {}).get("amount", 0)
    int_total = sb_total + dep_total + other_total + refund_int

    div_items = data.get("dividends", {}).get("items", [])
    div_total = sum(d.get("gross") or 0 for d in div_items)
    div_tds = sum(d.get("tds") or 0 for d in div_items)

    cg = data.get("capitalGains", {})
    lt_items = cg.get("longTerm", [])
    st_items = cg.get("shortTerm", [])
    for i in lt_items + st_items:
        if i.get("assetClass", "equity") != "equity":
            die_gap(f"non-equity capital-gains item '{i.get('name')}' ({i['assetClass']}): "
                    "s.112/slab-rate assets are not modelled")
    ltcg = sum(i["saleValue"] - i["cost"] for i in lt_items)
    loss_bf = cg.get("lossBroughtForward", 0)  # negative or 0
    setoff = min(ltcg, -loss_bf) if ltcg > 0 and loss_bf < 0 else 0
    net_ltcg = ltcg - setoff
    if net_ltcg < 0:
        die_gap(f"current-year net long-term capital LOSS Rs. {-net_ltcg}: s.70(3)/74 set-off and "
                "carry-forward of a fresh LT loss is not modelled (it must NOT reduce other income)")

    # Special-rate legs, split at the FA (No. 2) 2024 rate-change date (23-07-2024).
    # Every FY 2025-26 (AY 2026-27) transfer is on/after; undated items land there too.
    RATE_CHANGE = (2024, 7, 23)
    def _legs(items):
        before = sum(i["saleValue"] - i["cost"] for i in items
                     if _sale_date_key(i) != (0, 0, 0) and _sale_date_key(i) < RATE_CHANGE)
        total = sum(i["saleValue"] - i["cost"] for i in items)
        return before, total - before

    st_before, st_after = _legs(st_items)
    lt_before, lt_after = _legs(lt_items)
    if lt_before <= 0 or lt_after <= 0:
        # one leg zero/negative: the whole net gain sits in the other leg
        lt_before_net = max(0, net_ltcg) if lt_after <= 0 else 0
        lt_after_net = max(0, net_ltcg) if lt_after > 0 else 0
    else:
        # brought-forward loss set off against the before-leg first (chronological)
        lt_before_net = max(0, lt_before - setoff)
        lt_after_net = lt_after - max(0, setoff - lt_before)

    stcg_gross = sum(i["saleValue"] - i["cost"] for i in st_items)
    if stcg_gross < 0:
        die_gap(f"net short-term capital LOSS Rs. {-stcg_gross}: ST-loss set-off ordering "
                "(against LTCG / carry-forward) is not modelled")
    if st_before <= 0 or st_after <= 0:
        # one leg zero/negative: intra-year netting leaves the gain in the other leg
        st_before_net = max(0, stcg_gross) if st_after <= 0 else 0
        st_after_net = max(0, stcg_gross) if st_after > 0 else 0
    else:
        st_before_net, st_after_net = st_before, st_after
    special_gross = stcg_gross + net_ltcg

    hp = data.get("houseProperty", {})
    hp_props = hp.get("properties", [])
    hp_calc = []
    for p in hp_props:
        alv = p["annualLetableValue"]
        tax = p.get("propertyTax", 0)
        balance_alv = alv - tax
        share_pct = p.get("assesseeSharePercent", 100)
        owned_share = round(balance_alv * share_pct / 100)
        thirty_pct = round(owned_share * 0.3)
        income_hp = owned_share - thirty_pct
        hp_calc.append({"alv": alv, "tax": tax, "balance_alv": balance_alv, "share_pct": share_pct,
                         "owned_share": owned_share, "thirty_pct": thirty_pct, "income_hp": income_hp,
                         "tds": p.get("tds", 0)})
    hp_income_total = sum(c["income_hp"] for c in hp_calc)
    hp_tds_total = sum(c["tds"] for c in hp_calc)

    os_total = int_total + div_total
    gti = salary_gross + presumptive + net_ltcg + stcg_gross + os_total + hp_income_total
    tds_total = salary_tds + biz_tds + sb_tds + dep_tds + other_tds + div_tds + hp_tds_total
    ded_old = sum(d["amount"] for d in data.get("deductions", {}).get("oldRegime", [])) + std_ded_old
    ded_new = sum(d["amount"] for d in data.get("deductions", {}).get("newRegime", [])) + std_ded_new
    ti_old = gti - ded_old - special_gross
    ti_new = gti - ded_new - special_gross
    r_old, r_new = round(ti_old / 10) * 10, round(ti_new / 10) * 10

    # Full-liability engine computation per regime column (slab + 111A/112A special rates,
    # 112A exemption, resident-only basic-exemption adjustment and 87A gate).
    def _total(regime, slab_ti):
        return tax_cli("total", "--slab-income", str(slab_ti), "--regime", regime, "--ay", ay,
                       "--age", age, "--residential-status", res,
                       "--stcg-before", str(st_before_net), "--stcg-after", str(st_after_net),
                       "--ltcg-before", str(lt_before_net), "--ltcg-after", str(lt_after_net))
    comp_old = _total("old", r_old)
    comp_new = _total("new", r_new)
    if max(comp_old["totalIncome"], comp_new["totalIncome"]) > SURCHARGE_GUARD:
        die_gap(f"total income {max(comp_old['totalIncome'], comp_new['totalIncome'])} (incl. "
                "special-rate gains) exceeds Rs. 50 lakh: surcharge is not modelled in the sheet "
                "formulas")

    # ---- reconciliation targets (from the data file's TIS/26AS block)
    rec = data.get("reconciliation", {})
    checks = [
        ("salary vs TIS", salary_gross or None, rec.get("tisSalary")),
        ("dividends vs TIS", div_total, rec.get("tisDividend")),
        ("dividend TDS vs 26AS", div_tds, rec.get("tisDividendTds")),
        ("savings interest vs TIS", sb_total, rec.get("tisSavingsInterest")),
        ("savings interest TDS vs 26AS", sb_tds or None, rec.get("tisSavingsInterestTds")),
        ("deposit interest vs TIS", dep_total, rec.get("tisDepositInterest")),
        ("other interest vs TIS", other_total, rec.get("tisOtherInterest")),
        ("business receipts vs TIS", biz_gross, rec.get("tisBusinessReceipts")),
        ("securities sale vs TIS",
         (sum(i["saleValue"] for i in lt_items) + sum(i["saleValue"] for i in st_items)) or None,
         rec.get("tisSecuritiesSale")),
        ("total TDS vs 26AS", tds_total, rec.get("totalTds")),
        ("house property rent (100% ALV) vs 26AS",
         sum(c["alv"] for c in hp_calc) or None, rec.get("tisRent")),
        ("house property TDS vs 26AS", hp_tds_total or None, rec.get("tisRentTds")),
    ]
    for label, got, want in checks:
        if want is not None and got != want:
            die(f"reconciliation failed - {label}: data sums to {got}, target says {want}. "
                "Fix the DATA (re-read the source figure), never the check - and never invent a "
                "number to make it pass. If the same check fails twice for the same reason, stop "
                "guessing: re-run bun run extract (or re-read the PDF) and compare.")

    wb = openpyxl.Workbook()

    # ================= supporting sheets first (income sheet references them)
    div_cells = build_dividends_sheet(wb, data, ay, prior) if div_items else None
    int_cells = build_interest_sheet(wb, data, ay, prior)
    cg_cells = build_capital_gains_sheet(wb, data, ay, setoff) if (lt_items or loss_bf or st_items) else None
    if cg.get("lossHistory") or loss_bf:
        build_loss_sheet(wb, data, ay, setoff)

    # ================= income sheet
    ws = wb.create_sheet(f"Income_AY {short}", 0)
    wb.remove(wb["Sheet"])
    s = Sheet(ws)
    s.widths({"A": 3, "B": 38, "C": 6, "D": 30, "E": 14, "F": 16, "G": 14, "H": 14,
              "I": 13, "J": 3, "K": 16, "L": 3, "M": 13, "N": 11, "O": 7, "P": 7})

    s.put("C1", client["name"], NAME_FONT); s.r = 2
    s.row({"B": client.get("address", "")})
    s.row({"B": f"Aadhar Card No. {client.get('aadhaar', '')}"})
    s.row({"B": "Assessment Year :", "D": ay, "F": "Prevs.Yr.Ending :",
           "G": f"31/03/{ay[:4]}", "H": f"{client.get('itrForm', '')} to be filed"})
    regime_notes = client.get("regimeNotes", [])
    s.row({"B": "Status          :", "D": client.get("status", "01-Individual"),
           "F": "Regime :", "G": "NTR" if client.get("regime") == "new" else "OTR",
           "H": regime_notes[0] if regime_notes else ""})
    for n in regime_notes[1:]:
        s.row({"H": n}, note=True)
    s.row({"B": "Ward Number     :", "D": client.get("ward", ""),
           "F": "P.A.N. Number   :", "G": client["pan"],
           "H": client.get("portalPasswordNote", "")})
    s.row({"B": "Residntl.Status :", "D": client.get("residentialStatus", "Resident"),
           "F": "Date of Birth   :", "G": client["dob"],
           "H": {"lt60": "", "s60to79": "Senior citizen (60-79)", "gte80": "Super senior (80+)"}[age]})
    s.row({"B": "Return Status :", "D": client.get("returnStatus", "Original Return"),
           "F": "Gender:", "G": client.get("gender", "")})
    for b in client.get("banks", []):
        label = f"IFSC: {b['ifsc']}, A/c No. {b['account']}" + (" (For Refund)" if b.get("useForRefund") else "")
        s.row({"B": f"{b['label']}:", "D": label})
    s.skip()

    s.row({"H": f"AY {ay}", "I": f"AY {ay}", "M": f"AY {prior}", "N": f"AY {prior}"}, bold=True)
    s.row({"B": "Statement of Income:", "H": "Gross (Rs.)", "I": "TDS (Rs.)", "K": "TAN",
           "M": "Gross", "N": "TDS", "O": "26AS", "P": "AIS"}, bold=True)
    s.skip()

    # (1) salary
    r_sal = None
    if salary_gross:
        s.row({"B": "(1) Salary Income:"}, bold=True)
        r_sal = s.row({"B": salary.get("payer", ""), "H": salary_gross, "I": salary_tds,
                       "K": salary.get("tan", ""),
                       "M": salary.get("priorGross"), "N": salary.get("priorTds"),
                       "O": "ok" if salary.get("in26AS") else "", "P": "ok" if salary.get("inAIS") else ""})
        s.row({"B": f"Less: Standard Deduction u/s 16(ia) (Old Rs.{std_ded_old:,} / New Rs.{std_ded_new:,})"})
    else:
        s.row({"B": "(1) Salary Income:", "D": salary.get("note", "NIL"),
               "M": salary.get("priorNote", salary.get("prior", "NIL"))}, bold=True)
    s.skip()

    # (2) house property
    r_hp = None
    if hp_props:
        s.row({"B": "(2) Income from House property:"}, bold=True)
        hp_rows = []
        for p, c in zip(hp_props, hp_calc):
            tenant = p.get("tenant", {})
            co = p.get("coOwner")
            s.row({"B": p.get("address", "")})
            s.row({"B": "Tenant:", "D": tenant.get("name", ""), "K": tenant.get("tan", "")})
            if co:
                s.row({"B": "Co-owner:", "D": f"{co.get('name', '')} ({co.get('percentShare', 0)}%)"})
            r_alv = s.row({"B": "Annual Letable Value", "H": c["alv"]})
            r_tax = s.row({"B": "Less: Local Taxes (property tax)", "H": -c["tax"]})
            r_bal = s.row({"B": "Balance ALV", "H": f"=H{r_alv}+H{r_tax}"})
            r_owned = s.row({"B": f"Assessee's share ({c['share_pct']}%) of Balance ALV",
                             "H": f"=ROUND(H{r_bal}*{c['share_pct']}/100,0)"})
            r_30 = s.row({"B": "Less: 30% Standard Deduction", "H": f"=-ROUND(H{r_owned}*0.3,0)"})
            r_p_hp = s.row({"B": "Income chargeable under House Property",
                            "H": f"=H{r_owned}+H{r_30}", "I": c["tds"], "K": tenant.get("tan", ""),
                            "M": p.get("prior")}, bold=True)
            hp_rows.append(r_p_hp)
            s.skip()
        if len(hp_rows) > 1:
            r_hp = s.row({"B": "Total Income from House Property",
                          "H": "=" + "+".join(f"H{r}" for r in hp_rows),
                          "I": "=" + "+".join(f"I{r}" for r in hp_rows)}, bold=True)
        else:
            r_hp = hp_rows[0]
    else:
        s.row({"B": "(2) Income from House property:", "D": hp.get("note", "NIL"),
               "M": hp.get("prior", 0)}, bold=True)
        for line in hp.get("lines", []):
            s.row({"B": line})
    s.skip()

    # (3) capital gains
    r_ltcg = None
    s.row({"B": "(3) Capital Gains",
           **({"E": "Sale Value", "F": "Cost of Acqn."} if cg_cells else {})}, bold=True)
    if cg_cells:
        cgn = cg_cells["sheet"]
        r_ltcg = s.row({"B": f"LTCG (post 23/07/2024) - see {cgn} sheet",
                        "E": f"='{cgn}'!{cg_cells['sale']}",
                        "F": f"='{cgn}'!{cg_cells['cost']}",
                        "H": f"='{cgn}'!{cg_cells['gross']}", "M": cg.get("priorLtcg"), "P": "ok"})
        if setoff:
            s.row({"B": "Less: Set off against LT capital loss brought forward",
                   "H": f"='{cgn}'!{cg_cells['setoff']}", "K": f"b/f ({-loss_bf:,})"})
        r_net = s.row({"B": "Net Long Term Capital Gains", "H": f"='{cgn}'!{cg_cells['net']}",
                       "K": f"c/f ({-loss_bf - setoff:,}) - see LT loss sheet" if loss_bf else ""}, bold=True)
    else:
        r_net = s.row({"B": "Capital Gains", "D": "NIL", "H": 0, "M": cg.get("priorLtcg", "NIL")})
    r_stcg = None
    if cg_cells and cg_cells.get("stcgGross"):
        cgn = cg_cells["sheet"]
        r_stcg = s.row({"B": f"Short Term Capital Gains (u/s 111A) - see {cgn} sheet",
                        "E": f"='{cgn}'!{cg_cells['stcgSale']}",
                        "F": f"='{cgn}'!{cg_cells['stcgCost']}",
                        "H": f"='{cgn}'!{cg_cells['stcgGross']}", "M": cg.get("priorStcg")}, bold=True)
    else:
        s.row({"B": "Short Term Capital Gains", "D": cg.get("shortTermNote", "NIL"),
               "M": cg.get("priorStcg")})
    s.skip()

    # (4) business & profession
    s.row({"B": "(4) Income from Business & Profession:"}, bold=True)
    presumptive_rows = []
    for b in biz_items:
        pr = b.get("prior", {})
        rb = s.row({"B": b["payer"], "H": b["gross"], "I": b.get("tds", 0), "K": b.get("tan", ""),
                    "M": pr.get("gross"), "N": pr.get("tds"),
                    "O": "ok" if b.get("in26AS") else "", "P": "ok" if b.get("inAIS") else ""})
        rp = s.row({"B": b.get("notes", [""])[0],
                    "D": f"Income under Presumptive Tax u/s 44ADA ({b['presumptiveRate']:.0%})",
                    "H": f"=ROUND(H{rb}*{b['presumptiveRate']},0)", "M": pr.get("presumptive")})
        presumptive_rows.append((rb, rp))
        for n in b.get("notes", [])[1:]:
            s.row({"B": n}, note=True)
        for k, v in b.get("meta", {}).items():
            s.row({"B": f"{k}:", "D": v})
    if len(presumptive_rows) > 1:
        r_presum = s.row({"B": "Total presumptive income",
                          "H": "=" + "+".join(f"H{rp}" for _, rp in presumptive_rows)}, bold=True)
    else:
        r_presum = presumptive_rows[0][1] if presumptive_rows else s.row({"B": "None", "H": 0})
    biz_tds_rows = [rb for rb, _ in presumptive_rows]
    s.skip()

    # (5) other sources
    s.row({"B": "(5) Income from Other Sources:"}, bold=True)
    isn = int_cells["sheet"]
    r_dep = s.row({"B": "Interest on Term/Fixed Deposits", "H": f"='{isn}'!{int_cells['deposits']}",
                   "I": dep_tds, "K": (dep_items[0].get("tan", "") if dep_items else ""),
                   "M": (interest.get("depositsPrior") or {}).get("gross"),
                   "N": (interest.get("depositsPrior") or {}).get("tds"),
                   "O": "ok" if dep_tds else "", "P": "ok"})
    r_sb = s.row({"B": "Savings Bank Accounts Interest", "H": f"='{isn}'!{int_cells['savings']}",
                  **({"I": sb_tds, "K": (sb_items[0].get("tan", "") if sb_items else ""),
                      "O": "ok"} if sb_tds else {}),
                  "M": interest.get("savingsPriorTotal"), "P": "ok"})
    r_bank = s.row({"B": "Total Interest from Banks", "H": f"=SUM(H{r_dep}:H{r_sb})",
                    "M": f"=SUM(M{r_dep}:M{r_sb})"}, bold=True)
    r_other = None
    if int_cells.get("other"):
        r_other = s.row({"B": "Interest on Securities / Other (Section 193)",
                         "H": f"='{isn}'!{int_cells['other']}",
                         "I": interest.get("otherTds", 0),
                         "O": "ok" if interest.get("otherTds") else "", "P": "ok"})
    ri = interest.get("refundInterest") or {}
    if ri:
        r_ri = s.row({"B": f"Interest on Income Tax Refund for AY {prior}", "H": ri.get("amount", 0),
                      "K": ri.get("note", ""), "M": ri.get("prior")})
        if ri.get("refundReceived"):
            s.row({"B": f"(Refund received {ri.get('receivedDate', '')}: Rs. {ri['refundReceived']:,} "
                        f"vs claimed Rs. {ri.get('refundClaimed', 0):,})"}, note=True)
    else:
        r_ri = None
    int_parts = [f"H{r_bank}"] + ([f"H{r_other}"] if r_other else []) + ([f"H{r_ri}"] if r_ri else [])
    r_int = s.row({"B": "Total Interest income", "H": "=" + "+".join(int_parts),
                   "M": f"=M{r_bank}" + (f"+M{r_ri}" if r_ri else "")}, bold=True)
    s.skip()
    dv = data.get("dividends", {})
    if div_cells:
        dvn = div_cells["sheet"]
        r_div = s.row({"B": f"Dividends (see {dvn} sheet)", "H": f"='{dvn}'!{div_cells['gross']}",
                       "I": f"='{dvn}'!{div_cells['tds']}",
                       "M": (dv.get("prior") or {}).get("total"), "N": (dv.get("prior") or {}).get("tds"),
                       "O": "ok" if div_tds else "", "P": "ok"})
    else:
        r_div = s.row({"B": "Dividends", "D": "NIL", "H": 0})
    s.skip()
    os_tds_parts = [f"I{r_dep}"] + ([f"I{r_sb}"] if sb_tds else []) + \
                   ([f"I{r_other}"] if r_other else []) + [f"I{r_div}"]
    r_os = s.row({"B": "Total Income from Other Sources", "H": f"=H{r_int}+H{r_div}",
                  "I": "=" + "+".join(os_tds_parts),
                  "M": f"=M{r_int}+M{r_div}"}, bold=True)
    s.skip()

    py = data.get("priorYear", {})
    gti_h_parts = ([f"H{r_sal}"] if r_sal else []) + ([f"H{r_hp}"] if r_hp else []) + \
                  [f"H{r_presum}", f"H{r_net}"] + \
                  ([f"H{r_stcg}"] if r_stcg else []) + [f"H{r_os}"]
    gti_i_parts = ([f"I{r_sal}"] if r_sal else []) + ([f"I{r_hp}"] if r_hp else []) + \
                  [f"I{r}" for r in biz_tds_rows] + [f"I{r_os}"]
    r_gti = s.row({"B": "Gross Total Income", "H": "=" + "+".join(gti_h_parts),
                   "I": "=" + "+".join(gti_i_parts),
                   "K": "Total TDS ->", "M": py.get("gti"), "N": py.get("tds")}, bold=True)
    s.skip()

    # deductions
    ded = data.get("deductions", {})
    s.row({"B": "Less: Deductions under Chapter VIA:", "G": "Old Tax Regime",
           "I": "New Tax Regime"}, bold=True)
    ded_first = s.r
    if salary_gross:
        s.row({"B": "Standard Deduction u/s 16(ia)", "G": std_ded_old, "I": std_ded_new})
    for d in ded.get("oldRegime", []):
        s.row({"B": d["section"], "D": d.get("desc", ""), "G": d["amount"], "I": 0})
    for d in ded.get("newRegime", []):
        s.row({"B": d["section"], "D": d.get("desc", ""), "I": d["amount"]})
    ded_last = s.r - 1
    if ded_last >= ded_first:
        r_ded = s.row({"B": "Total Deductions", "G": f"=SUM(G{ded_first}:G{ded_last})",
                       "I": f"=SUM(I{ded_first}:I{ded_last})"}, bold=True)
    else:
        r_ded = s.row({"B": "Total Deductions", "G": 0, "I": 0}, bold=True)
    if ded.get("note"):
        s.row({"B": ded["note"]}, note=True)
    s.skip()

    # tax computation: G = old regime, I = new regime
    s.row({"G": "Old Tax Regime", "I": "New Tax Regime"}, bold=True)
    r_ti = s.row({"B": "Taxable Income (slab-rate income only; excl. special-rate capital gains)",
                  "G": f"=H{r_gti}-G{r_ded}-{special_gross}", "I": f"=H{r_gti}-I{r_ded}-{special_gross}"},
                 bold=True)
    r_round = s.row({"B": "Rounded u/s 288A", "G": f"=ROUND(G{r_ti}/10,0)*10",
                     "I": f"=ROUND(I{r_ti}/10,0)*10"})
    r_tax = s.row({"B": "Tax on Total Income (slab rates)",
                   "G": slab_formula(f"G{r_round}", rules_old["slabs"]),
                   "I": slab_formula(f"I{r_round}", rules_new["slabs"])})
    if res == "nri":
        r_reb = s.row({"B": "Less: Rebate u/s 87A (not available: non-resident)", "G": 0, "I": 0})
    else:
        # 87A eligibility looks at TOTAL income (incl. special-rate gains); the rebate
        # itself only ever offsets the slab-rate tax cell.
        r_reb = s.row({"B": "Less: Rebate u/s 87A (threshold incl. special-rate gains)"
                            if special_gross else "Less: Rebate u/s 87A",
                       "G": rebate_formula(f"(G{r_round}+{special_gross})", f"G{r_tax}", rules_old["rebate"]),
                       "I": rebate_formula(f"(I{r_round}+{special_gross})", f"I{r_tax}", rules_new["rebate"])})
    r_netx = s.row({"B": "Tax after rebate", "G": f"=G{r_tax}+G{r_reb}", "I": f"=I{r_tax}+I{r_reb}"})
    r_stcgtax = None
    if stcg_gross:
        r_stcgtax = s.row({"B": "Add: Tax on STCG u/s 111A (flat rate; no Chapter VI-A set-off)",
                           "G": comp_old["stcg111A"]["tax"], "I": comp_new["stcg111A"]["tax"]})
    r_ltcgtax = None
    if net_ltcg > 0:
        r_ltcgtax = s.row({"B": "Add: Tax on LTCG u/s 112A (above Rs.1,25,000 exemption)",
                           "G": comp_old["ltcg112A"]["tax"], "I": comp_new["ltcg112A"]["tax"]})
    r_reb2 = None
    if comp_old["rebate87AOnSpecial"] or comp_new["rebate87AOnSpecial"]:
        r_reb2 = s.row({"B": "Less: 87A rebate against 111A tax (old regime only)",
                        "G": -comp_old["rebate87AOnSpecial"], "I": -comp_new["rebate87AOnSpecial"]})
    if comp_old["basicExemptionAdjustment"] or comp_new["basicExemptionAdjustment"]:
        s.row({"B": f"(unexhausted basic exemption absorbs Rs.{comp_old['basicExemptionAdjustment']:,} "
                    f"old / Rs.{comp_new['basicExemptionAdjustment']:,} new of the gains above)"}, note=True)
    extra_rows = [r for r in (r_stcgtax, r_ltcgtax, r_reb2) if r]
    r_cess_base_g = f"G{r_netx}" + "".join(f"+G{r}" for r in extra_rows)
    r_cess_base_i = f"I{r_netx}" + "".join(f"+I{r}" for r in extra_rows)
    r_cess = s.row({"B": f"Add: Health & education cess ({rules_new['cessRate']:.0%})",
                    "G": f"=ROUND(({r_cess_base_g})*{rules_new['cessRate']},0)",
                    "I": f"=ROUND(({r_cess_base_i})*{rules_new['cessRate']},0)"})
    r_total_tax = s.row({"B": "Total Tax Payable",
                         "G": f"={r_cess_base_g}+G{r_cess}",
                         "I": f"={r_cess_base_i}+I{r_cess}"}, bold=True)
    s.row({"B": f"Slabs per scripts/lib/tax.ts for AY {ay} ({'senior' if age != 'lt60' else 'below 60'}"
               f"{'; NRI: no 87A, base-exemption slabs' if res == 'nri' else ''}); "
               "87A incl. marginal relief where applicable; 111A/112A flat rates per computeTotalTax. "
               "Surcharge not modelled (guarded at 50L total income)."}, note=True)
    s.skip()

    tp = data.get("taxesPaid", {})
    r_tds = s.row({"B": "Less: TDS", "G": f"=I{r_gti}", "I": f"=I{r_gti}"})
    r_adv = s.row({"B": "Less: Advance Tax", "G": tp.get("advance", 0), "I": tp.get("advance", 0)})
    r_sat = s.row({"B": "Less: Self Assessment Tax", "G": tp.get("selfAssessment", 0),
                   "I": tp.get("selfAssessment", 0)})
    r_refund = s.row({"B": "NET TAX REFUNDABLE / (PAYABLE)",
                      "G": f"=G{r_tds}+G{r_adv}+G{r_sat}-G{r_total_tax}",
                      "I": f"=I{r_tds}+I{r_adv}+I{r_sat}-I{r_total_tax}"}, bold=True)
    if data.get("refundBankNote"):
        s.row({"D": data["refundBankNote"]}, note=True)
    s.skip()

    exempt = data.get("exemptIncome", [])
    if exempt:
        s.row({"B": "Exempt Income:"}, bold=True)
        for e in exempt:
            s.row({"B": e["name"], "D": e.get("desc", ""), "H": e.get("amount"), "M": e.get("prior")})
        s.skip()

    notes = data.get("notes", [])
    if notes:
        s.row({"B": "Notes:"}, bold=True)
        for i, n in enumerate(notes):
            s.row({"B": f"{i + 1}. {n}"}, note=True)

    wb.save(out_path)

    # ---------------- verification: re-evaluate every formula written ----------------
    expected = {
        (ws.title, f"H{r_presum}"): presumptive,
        (ws.title, f"H{r_net}"): net_ltcg,
        (ws.title, f"H{r_int}"): int_total,
        (ws.title, f"H{r_os}"): os_total,
        (ws.title, f"H{r_gti}"): gti,
        (ws.title, f"I{r_gti}"): tds_total,
        (ws.title, f"G{r_ded}"): ded_old,
        (ws.title, f"I{r_ded}"): ded_new,
        (ws.title, f"G{r_round}"): r_old,
        (ws.title, f"I{r_round}"): r_new,
        (ws.title, f"G{r_tax}"): py_band_tax(rules_old["slabs"], r_old),
        (ws.title, f"I{r_tax}"): py_band_tax(rules_new["slabs"], r_new),
    }
    if div_cells:
        expected[(div_cells["sheet"], div_cells["gross"])] = div_total
        expected[(div_cells["sheet"], div_cells["tds"])] = div_tds
    if cg_cells and r_ltcg is not None and lt_items:
        expected[(ws.title, f"E{r_ltcg}")] = sum(i["saleValue"] for i in lt_items)
        expected[(ws.title, f"F{r_ltcg}")] = sum(i["cost"] for i in lt_items)
    if cg_cells and r_stcg is not None and st_items:
        expected[(ws.title, f"E{r_stcg}")] = sum(i["saleValue"] for i in st_items)
        expected[(ws.title, f"F{r_stcg}")] = sum(i["cost"] for i in st_items)
        expected[(ws.title, f"H{r_stcg}")] = stcg_gross
    if r_sal:
        expected[(ws.title, f"H{r_sal}")] = salary_gross
        expected[(ws.title, f"I{r_sal}")] = salary_tds
    if r_hp:
        expected[(ws.title, f"H{r_hp}")] = hp_income_total
        expected[(ws.title, f"I{r_hp}")] = hp_tds_total
    expected[(int_cells["sheet"], int_cells["savings"])] = sb_total
    expected[(int_cells["sheet"], int_cells["deposits"])] = dep_total

    values = verify(out_path, expected)

    # cross-check the sheet's tax result against the engine's full-liability computation
    # (slab + 111A/112A special rates + 87A gating; tolerance for 288A/B rounding)
    for regime, comp in [("old", comp_old), ("new", comp_new)]:
        col = "G" if regime == "old" else "I"
        sheet_total = values[(ws.title, f"{col}{r_total_tax}")]
        if abs(sheet_total - comp["total"]) > 10:
            die(f"{regime}-regime tax mismatch: sheet {sheet_total} vs engine {comp['total']}")
        sheet_refund = values[(ws.title, f"{col}{r_refund}")]
        print(f"  {regime} regime: tax {round(sheet_total):,} -> refundable {round(sheet_refund):,}")

    print(f"verified {len(expected)} anchor cells + engine cross-check - all reconcile")

    # ---- carry-forward sidecar: what next year's `bun run extract` needs but the
    # data file doesn't carry (computed GTI, the loss set-off actually applied).
    # Written only after full verification, so its figures are always reconciled.
    cf_next = loss_bf + setoff  # negative or 0; setoff moves it toward 0
    sidecar = {
        "_generated": "by build-statement.py after a verified build - consumed by next year's bun run extract; do not hand-edit",
        "assessmentYear": ay,
        "regime": client.get("regime"),
        "gti": round(gti),
        "totalIncome": {"old": round(comp_old["totalIncome"]), "new": round(comp_new["totalIncome"])},
        "tds": round(tds_total),
        "capitalLoss": {
            "broughtForward": loss_bf,
            "setOffThisYear": setoff,
            "carryForwardNext": cf_next,
            "lossHistoryAppend": (
                {"ay": f"AY {ay}", "particulars": "Set off against LTCG", "movement": setoff}
                if setoff else None
            ),
        },
    }
    # legacy clients keep the Excel in an AY subfolder; the sidecar belongs next to
    # the data file so extract finds it - prefer the data file's directory
    base_dir = Path(data_dir) if data_dir else Path(out_path).parent
    sidecar_path = base_dir / f"AY{ay}-carryforward.json"
    sidecar_path.write_text(json.dumps(sidecar, indent=2) + "\n")
    print(f"carry-forward sidecar: {sidecar_path}")
    return out_path


def build_dividends_sheet(wb, data, ay, prior):
    dv = data["dividends"]
    ws = wb.create_sheet("Dividends")
    s = Sheet(ws)
    s.widths({"A": 42, "D": 12, "E": 10, "F": 7, "G": 13, "H": 12, "J": 10,
              "L": 11, "M": 11, "N": 11, "O": 11, "P": 11, "Q": 7, "R": 7})
    s.put("A1", data["client"]["name"], NAME_FONT)
    s.put("A2", f"AY {ay}")
    s.r = 4
    s.row({"A": "Statement of Dividends Income (Taxable)", "D": f"AY {ay}", "H": f"AY {prior}",
           "L": f"Period of receipt (AY {ay})"}, bold=True)
    s.row({"D": "Gross (Rs.)", "E": "TDS (Rs.)", "F": "Src", "G": "TAN", "H": "Gross",
           "J": "TDS", "Q": "26AS", "R": "AIS"}, bold=True)
    s.row({"L": "up to 15/6", "M": "16/6 to 15/9", "N": "16/9 to 15/12", "O": "16/12 to 15/3",
           "P": "16/3 to 31/3"}, bold=True)
    first = s.r
    for d in dv["items"]:
        cells = {"A": d["name"]}
        if d.get("gross") is not None:
            cells.update({"D": d["gross"], "R": "ok"})
            if d.get("tds"):
                cells.update({"E": d["tds"], "Q": "ok"})
            else:
                cells["F"] = "AIS"
        if d.get("tan"):
            cells["G"] = d["tan"]
        if d.get("priorGross") is not None:
            cells["H"] = d["priorGross"]
        if d.get("priorTds") is not None:
            cells["J"] = d["priorTds"]
        for col, v in (d.get("periods") or {}).items():
            cells[col] = v
        s.row(cells)
    last = s.r - 1
    t = s.r + 1
    s.r = t
    s.row({"A": "Total Dividends",
           **{c: f"=SUM({c}{first}:{c}{last})" for c in "DEHJLMNOP"}}, bold=True)
    for n in dv.get("notes", []):
        s.row({"A": n}, note=True)
    return {"sheet": ws.title, "gross": f"D{t}", "tds": f"E{t}"}


def build_interest_sheet(wb, data, ay, prior):
    interest = data.get("interest", {})
    ws = wb.create_sheet("Interest income")
    s = Sheet(ws)
    s.widths({"A": 30, "B": 9, "C": 20, "E": 13, "F": 12, "G": 10, "H": 8,
              "J": 12, "K": 12, "L": 44, "M": 7, "N": 7})
    s.put("A1", data["client"]["displayName"], NAME_FONT)
    s.put("A2", f"AY {ay}")
    s.r = 3
    s.row({"F": f"AY {ay}", "K": f"AY {prior}"}, bold=True)
    s.row({"A": "Savings Bank Accounts Interest", "F": "Gross", "G": "TDS", "H": "Source",
           "K": "Gross", "M": "26AS", "N": "AIS"}, bold=True)
    sb_rows = []
    for i in interest.get("savings", []):
        sb_rows.append(s.row({"A": i["bank"], "B": "A/c No.", "C": i.get("account", ""),
                              "F": i["amount"], "H": i.get("source", "AIS"),
                              "K": i.get("prior"), "N": "ok"}))
        if i.get("ifsc"):
            s.row({"B": "IFSC Code:", "C": i["ifsc"]})
    r_sb = s.row({"A": "Total - SB Accounts interest",
                  "F": "=" + "+".join(f"F{r}" for r in sb_rows) if sb_rows else 0,
                  "K": interest.get("savingsPriorTotal")}, bold=True)
    s.skip()
    s.row({"A": "Term / Fixed Deposit interest"}, bold=True)
    dep_rows = []
    for i in interest.get("deposits", []):
        dep_rows.append(s.row({"A": i["bank"], "B": "A/c No.", "C": i.get("account", ""),
                               "E": i.get("tan", ""), "F": i["amount"],
                               "H": i.get("source", "AIS"), "N": "ok"}))
    depp = interest.get("depositsPrior") or {}
    r_dep = s.row({"A": "Total TD / FD interest",
                   "F": "=" + "+".join(f"F{r}" for r in dep_rows) if dep_rows else 0,
                   "G": interest.get("depositsTds", 0),
                   "K": depp.get("gross"), "L": depp.get("tds"), "M": "ok"}, bold=True)
    if interest.get("depositsNote"):
        s.row({"A": interest["depositsNote"]}, note=True)
    s.skip()
    other_items = interest.get("other", [])
    r_other = None
    if other_items:
        s.row({"A": "Interest on Securities / Other (Section 193)"}, bold=True)
        other_rows = []
        for i in other_items:
            other_rows.append(s.row({"A": i["payer"], "E": i.get("tan", ""), "F": i["amount"],
                                     "H": i.get("source", "AIS"), "N": "ok"}))
        r_other = s.row({"A": "Total - Interest on Securities / Other",
                         "F": "=" + "+".join(f"F{r}" for r in other_rows),
                         "G": interest.get("otherTds", 0)}, bold=True)
        if interest.get("otherNote"):
            s.row({"A": interest["otherNote"]}, note=True)
        s.skip()
    ri = interest.get("refundInterest") or {}
    r_ri = None
    if ri:
        s.row({"A": "Interest on Income Tax Refund", "E": f"AY {prior}",
               "J": "refund Rs.", "K": "interest Rs."}, bold=True)
        r_ri = s.row({"A": f"Refund received {ri.get('receivedDate', '')} (ECS)",
                      "F": ri.get("amount", 0), "H": "est.", "J": ri.get("refundReceived"),
                      "L": ri.get("note", "")})
        s.skip()
    total = f"=F{r_sb}+F{r_dep}" + (f"+F{r_other}" if r_other else "") + (f"+F{r_ri}" if r_ri else "")
    s.row({"A": "Total Interest income", "F": total, "G": f"=G{r_dep}",
           "K": interest.get("priorTotal"), "L": interest.get("priorTds")}, bold=True)
    return {"sheet": ws.title, "savings": f"F{r_sb}", "deposits": f"F{r_dep}",
            "other": f"F{r_other}" if r_other else None}


def _sale_date_key(item):
    d = item.get("saleDate", "")
    try:
        dd, mm, yyyy = d.split("/")
        return (int(yyyy), int(mm), int(dd))
    except ValueError:
        return (0, 0, 0)


def build_capital_gains_sheet(wb, data, ay, setoff):
    cg = data["capitalGains"]
    ws = wb.create_sheet("Capital Gains")
    s = Sheet(ws)
    s.widths({"A": 40, "B": 15, "C": 12, "D": 12, "F": 16, "H": 12, "J": 12})
    s.put("A2", data["client"]["displayName"], NAME_FONT)
    s.put("A4", f"AY {ay}")
    s.r = 6
    s.row({"A": "Capital Gains", "B": "ISIN", "C": "Date of Sale", "D": "Sale Value",
           "F": "Cost of Acquisition", "H": "LTCG", "J": "Source"}, bold=True)
    s.row({"A": "Long Term (post 23/07/2024)"}, bold=True)
    first = s.r
    for i in sorted(cg.get("longTerm", []), key=_sale_date_key):
        source = i.get("source", "")
        if i.get("costUnconfirmed"):
            source = (source + " - UNCONFIRMED COST" if source else "UNCONFIRMED COST")
        r = s.row({"A": i["name"], "B": i.get("isin", ""), "C": i.get("saleDate", ""),
                   "D": i["saleValue"], "F": i["cost"], "J": source})
        s.put(f"H{r}", f"=D{r}-F{r}", BASE, ACC)
        if i.get("costUnconfirmed"):
            s.row({"A": "  cost per AIS depository data - confirm against broker contract note"}, note=True)
    last = s.r - 1
    if last >= first:
        r_gross = s.row({"A": "Total LTCG", "D": f"=SUM(D{first}:D{last})", "F": f"=SUM(F{first}:F{last})",
                         "H": f"=SUM(H{first}:H{last})"}, bold=True)
    else:
        r_gross = s.row({"A": "Total LTCG", "D": 0, "F": 0, "H": 0}, bold=True)
    s.skip()
    r_setoff = s.row({"A": "Less: Set off against LT capital loss b/f", "H": -setoff, "J": "Sch BFLA"})
    r_netr = s.row({"A": "Net taxable LTCG", "H": f"=H{r_gross}+H{r_setoff}"}, bold=True)
    s.skip()

    st_items = cg.get("shortTerm", [])
    r_stgross = None
    if st_items:
        s.row({"A": "Short Term (post 23/07/2024) - taxable u/s 111A"}, bold=True)
        st_first = s.r
        for i in sorted(st_items, key=_sale_date_key):
            source = i.get("source", "")
            if i.get("costUnconfirmed"):
                source = (source + " - UNCONFIRMED COST" if source else "UNCONFIRMED COST")
            r = s.row({"A": i["name"], "B": i.get("isin", ""), "C": i.get("saleDate", ""),
                       "D": i["saleValue"], "F": i["cost"], "J": source})
            s.put(f"H{r}", f"=D{r}-F{r}", BASE, ACC)
            if i.get("costUnconfirmed"):
                s.row({"A": "  cost per AIS depository data - confirm against broker contract note"}, note=True)
        st_last = s.r - 1
        r_stgross = s.row({"A": "Total STCG (u/s 111A)", "D": f"=SUM(D{st_first}:D{st_last})",
                           "F": f"=SUM(F{st_first}:F{st_last})", "H": f"=SUM(H{st_first}:H{st_last})"},
                          bold=True)
        s.skip()
    else:
        s.row({"A": f"Short Term: {cg.get('shortTermNote', 'NIL')}"})

    for n in cg.get("notes", []):
        s.row({"A": n}, note=True)
    return {"sheet": ws.title, "gross": f"H{r_gross}", "sale": f"D{r_gross}",
            "cost": f"F{r_gross}", "setoff": f"H{r_setoff}", "net": f"H{r_netr}",
            "stcgGross": f"H{r_stgross}" if r_stgross else None,
            "stcgSale": f"D{r_stgross}" if r_stgross else None,
            "stcgCost": f"F{r_stgross}" if r_stgross else None}


def build_loss_sheet(wb, data, ay, setoff):
    cg = data["capitalGains"]
    ws = wb.create_sheet("LT Capital Loss c-f")
    s = Sheet(ws)
    s.widths({"A": 22, "B": 52, "F": 14, "H": 16})
    s.put("A1", data["client"]["displayName"], NAME_FONT)
    s.put("A2", "Long Term Capital Loss - roll-forward", BOLD)
    s.r = 4
    s.row({"A": "AY", "B": "Particulars", "F": "Movement (Rs.)", "H": "Balance c/f (Rs.)"}, bold=True)
    prev = None
    for h in cg.get("lossHistory", []):
        r = s.row({"A": h["ay"], "B": h["particulars"], "F": h["movement"],
                   "H": f"=H{prev}+F{s.r}" if prev else f"=F{s.r}"})
        prev = r
    if setoff:
        prev = s.row({"A": f"AY {ay}", "B": "Set off against LTCG (current year)", "F": setoff,
                      "H": f"=H{prev}+F{s.r}" if prev else f"=F{s.r}"})
    s.skip()
    next_ay = f"{int(ay[:4]) + 1}-{int(ay[5:]) + 1}"
    s.row({"A": f"Balance to carry forward to AY {next_ay}", "H": f"=H{prev}" if prev else 0}, bold=True)
    s.skip()
    for n in cg.get("lossNotes", []):
        s.row({"A": n}, note=True)


# ---------------------------------------------------------------- verifier
def verify(path, expected):
    """Re-evaluate every formula in the saved workbook and assert anchors."""
    wb = openpyxl.load_workbook(path)
    env, formulas = {}, {}
    for sh in wb.worksheets:
        for row in sh.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    formulas[(sh.title, c.coordinate)] = c.value
                elif isinstance(c.value, (int, float)):
                    env[(sh.title, c.coordinate)] = c.value

    def _if(cond, a, b):
        return a if cond else b

    def cellval(sheet, coord):
        key = (sheet, coord)
        if key in env:
            return env[key]
        if key in formulas:
            env[key] = evaluate(sheet, formulas[key])
            return env[key]
        return 0

    def evaluate(sheet, formula):
        expr = formula[1:].replace("$", "")
        expr = re.sub(r"'([^']+)'!([A-Z]{1,3}\d+)",
                      lambda m: repr(cellval(m.group(1), m.group(2))), expr)
        expr = re.sub(r"([A-Za-z][A-Za-z0-9 ]*?)!([A-Z]{1,3}\d+)",
                      lambda m: repr(cellval(m.group(1), m.group(2))), expr)

        def sumrange(m):
            col, r1, r2 = m.group(1), int(m.group(2)), int(m.group(4))
            assert col == m.group(3), f"non-column SUM in {formula}"
            return repr(sum(cellval(sheet, f"{col}{r}") for r in range(r1, r2 + 1)))

        expr = re.sub(r"SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)", sumrange, expr)
        expr = expr.replace("IF(", "_if(").replace("MIN(", "min(").replace("MAX(", "max(")
        expr = expr.replace("ROUND(", "round(")
        expr = re.sub(r"\b([A-Z]{1,3}\d+)\b", lambda m: repr(cellval(sheet, m.group(1))), expr)
        return eval(expr, {"_if": _if, "min": min, "max": max, "round": round})

    bad = []
    for (sheet, coord), want in expected.items():
        got = cellval(sheet, coord)
        if abs(got - want) > 0.01:
            bad.append(f"{sheet}!{coord}: sheet says {got}, expected {want}")
    if bad:
        die("verification failed:\n  " + "\n  ".join(bad))
    for key in list(formulas):  # force-evaluate everything to catch broken refs
        cellval(*key)
    return env


def main():
    args = [a for a in sys.argv[1:]]
    out = None
    if "--out" in args:
        i = args.index("--out")
        out = args[i + 1]
        del args[i:i + 2]
    if len(args) != 1:
        die("usage: python3 scripts/build-statement.py <statement-data.json> [--out <file.xlsx>]")
    data_path = Path(args[0])
    data = json.loads(data_path.read_text())
    gaps = data.get("_gaps") or []
    if gaps:
        die("unresolved extraction gaps in the data file - resolve each one and delete its _gaps "
            "entry (never build over a gap):\n  "
            + "\n  ".join(f"[{g.get('field')}] {g.get('action')}" for g in gaps))
    if not out:
        name = f"{data['client']['displayName']}_AY {ay_short(data['assessmentYear'])}_Statement.xlsx"
        out = data_path.parent / name
    result = build(data, str(out), data_dir=data_path.parent)
    print(f"saved: {result}")


if __name__ == "__main__":
    main()
